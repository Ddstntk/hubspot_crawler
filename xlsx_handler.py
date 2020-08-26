import openpyxl
from openpyxl import load_workbook

# workbook = load_workbook('./hubspot_list.xlsx')
# worksheet = workbook.get_sheet_by_name('Meta Data')

# for row_cells in worksheet.iter_rows():
#     print('\n')
#     for cell in row_cells:
#         print('%s: cell.value=%s' % (cell, cell.value))
    
class XlsHandler:
    # readUrls('./hubspot_list.xlsx','Meta Data')
    def __init__(self, filename, sheetname):
        self.url_list = []
        self.readUrls(filename, sheetname)

    def readUrls(self, filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=3):
            self.url_list.append(row[0].value)

    def getUrls(self):
        return self.url_list
    # readUrls('./hubspot_list.xlsx','Meta Data')


xd = XlsHandler('./hubspot_list.xlsx','Meta Data')
print(xd.getUrls())